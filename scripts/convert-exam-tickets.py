#!/usr/bin/env python3
"""
Синхронизация экзаменационных билетов с публичной ссылкой Яндекс Диска.

Источник: https://disk.yandex.ru/i/oAuOyVb4OkmNtA
          (файл «Экзаминационные билеты_app.xlsx»)

Скрипт работает по тому же принципу, что и scripts/sync-devices.py
(раздел «Приборы»):

  1. Через Yandex Disk Public API (https://cloud-api.yandex.net/v1/disk/public/resources)
     получает download_url по публичной ссылке.
  2. Скачивает XLSX-файл напрямую с Яндекс Диска.
  3. Парсит 4 листа («4 разряд», «5 разряд», «6 разряд», «До 1000 В»).
  4. Сохраняет результат в data/exam-tickets.json.

Переменные окружения:
  EXAM_TICKETS_PUBLIC_KEY — публичная ссылка
      (по умолчанию https://disk.yandex.ru/i/oAuOyVb4OkmNtA)

Если нет интернета или API недоступен — использует уже существующий
data/exam-tickets.json как заглушку (PWA продолжает работать с последними
закоммиченными данными).

Полностью заменяет прежнюю реализацию на OneDrive-ссылке, которая тянула
файл из основного репозитория kip8. Теперь источник независимый и хостится
на Яндекс Диске — так же, как и приборы.
"""

import json
import os
import re
import sys
from pathlib import Path

import requests
import openpyxl


# ============================================================
# Настройки Яндекс Диска
# ============================================================
YANDEX_PUBLIC_API = 'https://cloud-api.yandex.net/v1/disk/public/resources'
DEFAULT_PUBLIC_KEY = 'https://disk.yandex.ru/i/oAuOyVb4OkmNtA'

DOWNLOAD_DIR = Path('/tmp/exam_tickets_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'exam-tickets.json'


# ============================================================
# Конфигурация листов
# ============================================================
SHEETS_CONFIG = {
    "4 разряд": {"id": "tickets-4", "title": "Билеты на 4 разряд"},
    "5 разряд": {"id": "tickets-5", "title": "Билеты на 5 разряд"},
    "6 разряд": {"id": "tickets-6", "title": "Билеты на 6 разряд"},
    "До 1000 В": {"id": "tickets-1000v", "title": "Билеты до 1000 В"},
}


# ============================================================
# Стандартизация имён полей
# ============================================================
# Excel-таблица содержит русские заголовки с пробелами и спецсимволами
# (№ билета, № вопроса, Название литературы и т.д.).
# Для JSON используем латинские snake_case имена — это упрощает
# работу в JavaScript, устраняет проблемы с кодировками в escape-
# последовательностях (\u2116 для № и т.д.) и делает код читаемее.
#
# Фронтенд PWA (index.html) читает новые имена, но также поддерживает
# старые русские имена как fallback — это обеспечивает плавную
# миграцию, если где-то остался старый JSON.
FIELD_NAME_MAP = {
    "ID": "id",
    "№ билета": "ticket_number",
    "№билета": "ticket_number",        # вариант без пробела (для надёжности)
    "№ вопроса": "question_number",
    "№вопроса": "question_number",     # вариант без пробела
    "Вопрос": "question",
    "Ответ": "answer",
    "Image": "image_url",
    "Название литературы": "literature_name",
    "Файл": "file_url",
}


def log(msg):
    print(f'[exam-tickets] {msg}', flush=True)


def _normalize_field_name(raw_name: str) -> str:
    """Конвертирует русское имя поля из Excel в стандартное латинское.

    Если поле не входит в FIELD_NAME_MAP — возвращает оригинальное имя
    (это позволяет добавлять новые столбцы в Excel без изменения кода).
    """
    if not raw_name:
        return ""
    name = raw_name.strip()
    return FIELD_NAME_MAP.get(name, name)


def _is_http_url(s: str) -> bool:
    """Возвращает True, если строка начинается с http:// или https://.

    Это ЕДИНСТВЕННЫЙ допустимый формат значения для поля Image.
    Любые другие значения (локальные пути, относительные пути) будут
    проигнорированы фронтендом PWA через функцию isWorkingUrl().
    """
    return s.strip().lower().startswith(("http://", "https://"))


def _normalize_image_field(raw_value: str) -> str:
    """Нормализует значение поля Image из Excel.

    Единый принцип: только HTTP/HTTPS-ссылки.
    - Если значение — URL (http/https) → пропускаем как есть.
    - Если значение — что-то другое (локальный путь Windows,
      относительный путь, имя файла) → возвращаем пустую строку,
      чтобы в JSON не попал неработающий путь.
    """
    if not raw_value or not raw_value.strip():
        return ""
    if _is_http_url(raw_value):
        return raw_value.strip()
    print(f"  [ВНИМАНИЕ] Поле Image содержит не-URL значение, "
          f"оно будет проигнорировано в PWA: {raw_value[:80]}",
          file=sys.stderr)
    return ""


def _normalize_file_field(raw_value: str) -> str:
    """Нормализует значение поля Файл (ссылка на файл литературы).

    Аналогично полю Image: принимаются только HTTP/HTTPS-ссылки.
    """
    if not raw_value or not raw_value.strip():
        return ""
    if _is_http_url(raw_value):
        return raw_value.strip()
    return ""


# ============================================================
# Скачивание через Yandex Disk Public API
# (по образцу scripts/sync-devices.py)
# ============================================================
def get_download_url(public_key):
    """Получает download_url для публичного файла через Yandex Disk Public API."""
    log(f'Запрос метаданных публичного файла: {public_key}')
    params = {'public_key': public_key}
    resp = requests.get(YANDEX_PUBLIC_API, params=params, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка API: HTTP {resp.status_code} — {resp.text[:200]}')
    data = resp.json()
    download_url = data.get('file')
    if not download_url:
        raise RuntimeError(
            f'Не удалось получить download_url: {json.dumps(data, ensure_ascii=False)[:500]}'
        )
    name = data.get('name', 'exam-tickets.xlsx')
    log(f'Имя файла на Яндекс Диске: {name}')
    return download_url, name


def download_file(url, filename):
    """Скачивает файл по URL."""
    local_path = DOWNLOAD_DIR / filename
    log(f'Скачивание: {url[:80]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code}')
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}'
        )
    return local_path


# ============================================================
# Парсинг XLSX → JSON
# ============================================================
def convert_xlsx_to_json(xlsx_path, json_path):
    """Конвертирует xlsx в JSON.

    Структура вывода (для обратной совместимости с фронтендом):
    {
      "tickets-4": {
        "title": "Билеты на 4 разряд",
        "sheet": "4 разряд",
        "headers": ["id", "ticket_number", ...],
        "rows": [ {...}, {...} ],
        "total": N
      },
      "tickets-5": {...},
      "tickets-6": {...},
      "tickets-1000v": {...}
    }
    """
    log(f'Парсинг {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    log(f'Листы в файле: {wb.sheetnames}')

    all_data = {}
    for sheet_name, config in SHEETS_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            log(f'Лист «{sheet_name}» не найден, пропускаю')
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h else "" for h in rows[0]]
        normalized_headers = [_normalize_field_name(h) for h in headers]
        data_rows = []

        for row in rows[1:]:
            obj = {}
            for i, val in enumerate(row):
                if i < len(normalized_headers):
                    field_name = normalized_headers[i]
                    cell_val = str(val) if val is not None else ""
                    # Единый принцип: только HTTP/HTTPS-ссылки.
                    if field_name == "image_url":
                        cell_val = _normalize_image_field(cell_val)
                    elif field_name == "file_url":
                        cell_val = _normalize_file_field(cell_val)
                    obj[field_name] = cell_val
            data_rows.append(obj)

        all_data[config["id"]] = {
            "title": config["title"],
            "sheet": sheet_name,
            "headers": normalized_headers,
            "rows": data_rows,
            "total": len(data_rows),
        }
        log(f'  {sheet_name}: {len(data_rows)} строк')

    json_path.parent.mkdir(parents=True, exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)
    size = json_path.stat().st_size
    log(f'JSON сохранён: {json_path} ({size / 1024:.1f} КБ)')
    return True


def main():
    public_key = os.environ.get('EXAM_TICKETS_PUBLIC_KEY', '').strip() or DEFAULT_PUBLIC_KEY

    try:
        # 1. Получить download_url
        download_url, filename = get_download_url(public_key)

        # 2. Скачать файл
        local_file = download_file(download_url, filename)

        # 3. Конвертировать в JSON
        if not convert_xlsx_to_json(local_file, JSON_OUT):
            return 1
        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())

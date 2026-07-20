#!/usr/bin/env python3
"""
Синхронизация блокировок с публичной ссылкой Яндекс Диска.

Источник: https://disk.yandex.ru/i/KNVTQ7Q6II7zbA
          (файл «Перечень КИП ИОС рабочий.xlsx», тот же что и для приборов)
Лист: "Блокировки_app"

Скрипт работает по тому же принципу, что и scripts/sync-devices.py
(раздел «Приборы»):

  1. Через Yandex Disk Public API (https://cloud-api.yandex.net/v1/disk/public/resources)
     получает download_url по публичной ссылке.
  2. Скачивает XLSX-файл напрямую с Яндекс Диска.
  3. Парсит лист "Блокировки_app" — заголовки в 1-й строке, данные со 2-й.
  4. Сохраняет результат в data/lockouts.json.

Переменные окружения:
  LOCKOUTS_PUBLIC_KEY — публичная ссылка
      (по умолчанию https://disk.yandex.ru/i/KNVTQ7Q6II7zbA)
  LOCKOUTS_SHEET_NAME — имя листа (по умолчанию "Блокировки_app")

Если нет интернета или API недоступен — используется уже существующий
data/lockouts.json как заглушка (PWA продолжает работать с последними
закоммиченными данными).
"""

import os
import sys
import json
import re
from pathlib import Path
from datetime import datetime

import requests
import openpyxl


# ============================================================
# Настройки Яндекс Диска
# ============================================================
YANDEX_PUBLIC_API = 'https://cloud-api.yandex.net/v1/disk/public/resources'
DEFAULT_PUBLIC_KEY = 'https://disk.yandex.ru/i/KNVTQ7Q6II7zbA'
DEFAULT_SHEET_NAME = 'Блокировки_app'

DOWNLOAD_DIR = Path('/tmp/lockouts_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'lockouts.json'


def log(msg):
    print(f'[lockouts] {msg}', flush=True)


# ============================================================
# Скачивание через Yandex Disk Public API
# (по образцу scripts/sync-devices.py)
# ============================================================
def get_download_url(public_key):
    """Получает download_url для публичного файла через Yandex Disk Public API."""
    log(f'Запрос метаданных публичного файла: {public_key}')
    params = {'public_key': public_key}
    resp = requests.get(YANDEX_PUBLIC_API, params=params, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка API: HTTP {resp.status_code} — {resp.text[:200]}')
    data = resp.json()
    download_url = data.get('file')
    if not download_url:
        raise RuntimeError(
            f'Не удалось получить download_url: {json.dumps(data, ensure_ascii=False)[:500]}'
        )
    name = data.get('name', 'devices.xlsx')
    log(f'Имя файла на Яндекс Диске: {name}')
    return download_url, name


def download_file(url, filename):
    """Скачивает файл по URL."""
    local_path = DOWNLOAD_DIR / filename
    log(f'Скачивание: {url[:80]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code}')
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}'
        )
    return local_path


# ============================================================
# Парсинг листа Блокировки_app → data/lockouts.json
# ============================================================
def parse_lockouts(xlsx_path, sheet_name):
    """
    Парсит лист sheet_name из XLSX-файла.
    Заголовки — в 1-й строке, данные — начиная со 2-й.
    Пропускает строки без ID и без Параметра.
    """
    log(f'Парсинг листа "{sheet_name}" из {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f'Лист "{sheet_name}" не найден. Доступные листы: {wb.sheetnames}'
        )

    ws = wb[sheet_name]
    log(f'Размер листа: {ws.max_row} строк × {ws.max_column} колонок')

    # Читаем заголовки из 1-й строки
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ''
        headers.append(val)
    log(f'Заголовки ({len(headers)}): {headers}')

    # Читаем данные
    lockouts = []
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            # Обработка дат (например, "Дата проверки")
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif val is not None:
                val = str(val).strip()
                # Убираем мягкие переносы и нормализуем пробелы
                val = val.replace('\xad', '').replace('\u00a0', ' ')
                val = re.sub(r'\s+', ' ', val).strip()
            else:
                val = ''
            row_values.append(val)

        # Создаём словарь "заголовок → значение"
        record = {}
        for h, v in zip(headers, row_values):
            if h:  # пропускаем пустые заголовки
                record[h] = v

        # Пропускаем строки без ID и без Параметра
        id_val = record.get('ID', '').strip() if isinstance(record.get('ID'), str) else record.get('ID')
        param_val = record.get('Параметр', '').strip() if isinstance(record.get('Параметр'), str) else record.get('Параметр')
        if (id_val == '' or id_val is None) and (param_val == '' or param_val is None):
            skipped += 1
            continue

        # Если ID — число, преобразуем
        if isinstance(id_val, str) and id_val.isdigit():
            record['ID'] = int(id_val)

        lockouts.append(record)

    log(f'Распарсено записей: {len(lockouts)}, пропущено: {skipped}')
    return lockouts, headers


def main():
    public_key = os.environ.get('LOCKOUTS_PUBLIC_KEY', '').strip() or DEFAULT_PUBLIC_KEY
    sheet_name = os.environ.get('LOCKOUTS_SHEET_NAME', '').strip() or DEFAULT_SHEET_NAME

    try:
        # 1. Получить download_url
        download_url, filename = get_download_url(public_key)

        # 2. Скачать файл
        local_file = download_file(download_url, filename)

        # 3. Распарсить лист
        lockouts, headers = parse_lockouts(local_file, sheet_name)

        # 4. Сохранить JSON
        out = {
            'title': 'Блокировки по производствам',
            'source': f'Yandex Disk (public): {public_key}',
            'sheet': sheet_name,
            'total_lockouts': len(lockouts),
            'headers': headers,
            'lockouts': lockouts,
        }
        JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
        with open(JSON_OUT, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        log(f'JSON сохранён: {JSON_OUT}')
        log(f'Всего блокировок: {len(lockouts)}')

        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())

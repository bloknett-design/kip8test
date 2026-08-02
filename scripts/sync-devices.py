#!/usr/bin/env python3
"""
Синхронизация перечня приборов КИП ИОС с Google Sheets.

Источник: https://docs.google.com/spreadsheets/d/1eUUwwulUvKUGWTgQ__XP-y7z1aEkt5Wy/edit
          (файл «Перечень КИП ИОС рабочий.xlsx», импортированный в Google Sheets)
Лист: "Приборы_app"

Скрипт:
1. Скачивает XLSX-экспорт напрямую из Google Sheets через export?format=xlsx.
   Google отдаёт файл без OAuth, если таблица доступна «у кого есть ссылка».
2. Парсит лист "Приборы_app" — заголовки в 1-й строке, данные со 2-й.
3. Для записей с Yandex Disk share-ссылками в поле «Изображение» — разрешает
   их через Yandex Disk Public API и заменяет на base64 data URI.
4. Сохраняет результат в data/devices.json.

Переменные окружения:
  DEVICES_SPREADSHEET_ID — ID Google Sheets
      (по умолчанию 1eUUwwulUvKUGWTgQ__XP-y7z1aEkt5Wy)
  DEVICES_SHEET_NAME — имя листа (по умолчанию "Приборы_app")
  DEVICES_GID — numeric ID листа (опционально; если задан, экспортирует
      конкретный лист через &gid=...). Если не задан — экспортируется вся книга.

Если нет интернета — используется уже существующий data/devices.json как заглушка.
"""
import os
import sys
import json
import re
import base64
import io
import time
from pathlib import Path
from datetime import datetime

import requests
import openpyxl
try:
    from PIL import Image
    HAS_PILLOW = True
except ImportError:
    HAS_PILLOW = False


# ============================================================
# Настройки Google Sheets
# ============================================================
DEFAULT_SPREADSHEET_ID = '1eUUwwulUvKUGWTgQ__XP-y7z1aEkt5Wy'
DEFAULT_SHEET_NAME = 'Приборы_app'

# Yandex Disk Public API — используется только для разрешения share-ссылок
# на картинки в поле «Изображение» (значения ячеек сохранены при импорте
# исходного xlsx в Google Sheets).
YANDEX_PUBLIC_API = 'https://cloud-api.yandex.net/v1/disk/public/resources'

DOWNLOAD_DIR = Path('/tmp/devices_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'devices.json'


def log(msg):
    print(f'[devices] {msg}', flush=True)


# ============================================================
# Скачивание XLSX напрямую из Google Sheets
# ============================================================
def download_file(spreadsheet_id, gid=None):
    """
    Скачивает XLSX-экспорт Google Sheets.

    URL: https://docs.google.com/spreadsheets/d/<ID>/export?format=xlsx[&gid=<GID>]
    Если gid не задан — экспортируется вся книга (все листы).
    """
    url = f'https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export?format=xlsx'
    if gid:
        url += f'&gid={gid}'

    log(f'Скачивание: {url[:100]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120, allow_redirects=True)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code} — {resp.text[:200]}')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}. '
            f'Возможно, таблица не опубликована или нет доступа.'
        )

    filename = 'devices.xlsx'
    local_path = DOWNLOAD_DIR / filename
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')
    return local_path


def parse_devices(xlsx_path, sheet_name):
    """
    Парсит лист sheet_name из XLSX-файла.
    Заголовки — в 1-й строке, данные — начиная со 2-й.
    Пропускает строки без ID или без Наименования.
    """
    log(f'Парсинг листа "{sheet_name}" из {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f'Лист "{sheet_name}" не найден. Доступные листы: {wb.sheetnames}')

    ws = wb[sheet_name]
    log(f'Размер листа: {ws.max_row} строк × {ws.max_column} колонок')

    # Читаем заголовки из 1-й строки
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ''
        headers.append(val)
    log(f'Заголовки ({len(headers)}): {headers}')

    # Читаем данные
    devices = []
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        row_values = []
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            # Обработка дат
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif val is not None:
                val = str(val).strip()
                # Убираем мягкие переносы и нормализуем пробелы
                val = val.replace('\xad', '').replace('\u00a0', ' ')
                val = re.sub(r'\s+', ' ', val).strip()
            else:
                val = ''
            row_values.append(val)

        # Создаём словарь "заголовок → значение"
        record = {}
        for h, v in zip(headers, row_values):
            if h:  # пропускаем пустые заголовки
                record[h] = v

        # Пропускаем строки без ID или без Наименования
        id_val = record.get('ID', '').strip()
        name_val = record.get('Наименование', '').strip()
        if not id_val and not name_val:
            skipped += 1
            continue

        # Если ID — число, преобразуем
        if id_val and id_val.isdigit():
            record['ID'] = int(id_val)

        devices.append(record)

    log(f'Распарсено записей: {len(devices)}, пропущено: {skipped}')
    return devices, headers


def resolve_share_link_images(devices, max_size=(150, 150)):
    """
    Для записей с share-ссылками (https://disk.yandex.ru/i/...) в поле 'Изображение':
    1. Разрешает ссылку через Yandex Disk API → file URL
    2. Скачивает картинку
    3. Уменьшает до max_size
    4. Заменяет share-ссылку на base64 data URI в поле 'Изображение'

    Записи с локальными путями или без картинки — не трогаются.

    Примечание: при импорте исходного xlsx в Google Sheets значения ячеек
    (включая Yandex share-ссылки) сохраняются, поэтому эта логика продолжает
    работать после переноса источника с Yandex Disk на Google Sheets.
    """
    if not HAS_PILLOW:
        log('Pillow не установлен — пропуск загрузки картинок')
        return devices

    # Соберём уникальные share-ссылки
    share_links = {}  # { shareLink: base64dataUri }
    for d in devices:
        img = (d.get('Изображение') or '').strip()
        if img.startswith('https://disk.yandex.ru/i/'):
            if img not in share_links:
                share_links[img] = None

    if not share_links:
        log('Share-ссылок не найдено — картинки не загружаются')
        return devices

    log(f'Найдено уникальных share-ссылок: {len(share_links)}')
    session = requests.Session()

    for i, link in enumerate(share_links.keys()):
        try:
            # 1. Получаем file URL через API
            log(f'  [{i+1}/{len(share_links)}] Разрешение: {link[:50]}...')
            api_resp = session.get(YANDEX_PUBLIC_API, params={
                'public_key': link,
            }, timeout=30)
            if api_resp.status_code != 200:
                log(f'    ✗ API HTTP {api_resp.status_code}')
                continue
            file_url = api_resp.json().get('file', '')
            if not file_url:
                log(f'    ✗ Нет file URL в ответе')
                continue

            # 2. Скачиваем картинку (в той же сессии — cookies сохраняются)
            img_resp = session.get(file_url, timeout=60)
            if img_resp.status_code != 200:
                log(f'    ✗ Download HTTP {img_resp.status_code}')
                continue

            # 3. Уменьшаем и конвертируем в base64
            img = Image.open(io.BytesIO(img_resp.content))
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            buf = io.BytesIO()
            if img.mode in ('RGBA', 'LA', 'P'):
                img.save(buf, format='PNG')
                mime = 'image/png'
            else:
                img = img.convert('RGB')
                img.save(buf, format='JPEG', quality=85)
                mime = 'image/jpeg'
            b64 = base64.b64encode(buf.getvalue()).decode('ascii')
            data_uri = f'data:{mime};base64,{b64}'
            share_links[link] = data_uri
            log(f'    ✓ {len(b64)/1024:.1f}KB')

        except Exception as e:
            log(f'    ✗ Ошибка: {e}')
        time.sleep(0.3)

    # Заменяем share-ссылки на base64 в записях
    replaced = 0
    for d in devices:
        img = (d.get('Изображение') or '').strip()
        if img in share_links and share_links[img]:
            d['Изображение'] = share_links[img]
            replaced += 1

    log(f'Заменено ссылок на base64: {replaced}')
    return devices


def main():
    spreadsheet_id = os.environ.get('DEVICES_SPREADSHEET_ID', '').strip() or DEFAULT_SPREADSHEET_ID
    sheet_name = os.environ.get('DEVICES_SHEET_NAME', '').strip() or DEFAULT_SHEET_NAME
    gid = os.environ.get('DEVICES_GID', '').strip() or None

    try:
        # 1. Скачать XLSX из Google Sheets
        local_file = download_file(spreadsheet_id, gid=gid)

        # 2. Распарсить лист
        devices, headers = parse_devices(local_file, sheet_name)

        # 3. Разрешить share-ссылки на картинки → base64
        devices = resolve_share_link_images(devices)

        # 4. Сохранить JSON
        out = {
            'title': 'Перечень приборов КИП ИОС',
            'source': f'Google Sheets: https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit',
            'sheet': sheet_name,
            'total_devices': len(devices),
            'headers': headers,
            'devices': devices,
        }
        JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
        with open(JSON_OUT, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        log(f'JSON сохранён: {JSON_OUT}')
        log(f'Всего приборов: {len(devices)}')

        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())

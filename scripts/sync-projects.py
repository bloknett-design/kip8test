#!/usr/bin/env python3
"""
Синхронизация проектов с публичной ссылкой Яндекс Диска.

Источник: https://disk.yandex.ru/i/hfIfJ3EaHMWp9A
          (файл «Перечень проектов КИП пр-ва ИОС.xlsx»)
Лист: "Проекты"

Скрипт работает по тому же принципу, что и scripts/sync-valves.py
(раздел «Клапаны»):

  1. Через Yandex Disk Public API (https://cloud-api.yandex.net/v1/disk/public/resources)
     получает download_url по публичной ссылке.
  2. Скачивает XLSX-файл напрямую с Яндекс Диска.
  3. Парсит лист "Проекты" — заголовки в 1-й строке, данные со 2-й.
     Примечание: заголовок «Отделение » в исходном файле имеет trailing space,
     нормализуем все заголовки (strip) — итоговый ключ «Отделение».
  4. Сохраняет результат в data/projects.json.

Переменные окружения:
  PROJECTS_PUBLIC_KEY — публичная ссылка
      (по умолчанию https://disk.yandex.ru/i/hfIfJ3EaHMWp9A)
  PROJECTS_SHEET_NAME — имя листа (по умолчанию "Проекты")

Если нет интернета или API недоступен — используется уже существующий
data/projects.json как заглушка (PWA продолжает работать с последними
закоммиченными данными).
"""

import os
import sys
import json
import re
from pathlib import Path
from datetime import datetime

import requests
import openpyxl


# ============================================================
# Настройки Яндекс Диска
# ============================================================
YANDEX_PUBLIC_API = 'https://cloud-api.yandex.net/v1/disk/public/resources'
DEFAULT_PUBLIC_KEY = 'https://disk.yandex.ru/i/hfIfJ3EaHMWp9A'
DEFAULT_SHEET_NAME = 'Проекты'

DOWNLOAD_DIR = Path('/tmp/projects_download')
DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)

PROJECT_ROOT = Path(__file__).resolve().parent.parent
JSON_OUT = PROJECT_ROOT / 'data' / 'projects.json'


def log(msg):
    print(f'[projects] {msg}', flush=True)


# ============================================================
# Скачивание через Yandex Disk Public API
# ============================================================
def get_download_url(public_key):
    """Получает download_url для публичного файла через Yandex Disk Public API."""
    log(f'Запрос метаданных публичного файла: {public_key}')
    params = {'public_key': public_key}
    resp = requests.get(YANDEX_PUBLIC_API, params=params, timeout=30)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка API: HTTP {resp.status_code} — {resp.text[:200]}')
    data = resp.json()
    download_url = data.get('file')
    if not download_url:
        raise RuntimeError(
            f'Не удалось получить download_url: {json.dumps(data, ensure_ascii=False)[:500]}'
        )
    name = data.get('name', 'projects.xlsx')
    log(f'Имя файла на Яндекс Диске: {name}')
    return download_url, name


def download_file(url, filename):
    """Скачивает файл по URL."""
    local_path = DOWNLOAD_DIR / filename
    log(f'Скачивание: {url[:80]}...')
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    resp = requests.get(url, headers=headers, timeout=120)
    if resp.status_code != 200:
        raise RuntimeError(f'Ошибка скачивания: HTTP {resp.status_code}')
    local_path.write_bytes(resp.content)
    file_size = local_path.stat().st_size
    log(f'Файл скачан: {local_path} ({file_size} байт)')

    # Проверяем, что это xlsx (ZIP, начинается с PK)
    if resp.content[:2] != b'PK':
        raise RuntimeError(
            f'Скачанный файл не является xlsx (не ZIP). '
            f'Первые байты: {resp.content[:4]!r}'
        )
    return local_path


# ============================================================
# Парсинг листа Проекты → data/projects.json
# ============================================================
def parse_projects(xlsx_path, sheet_name):
    """
    Парсит лист sheet_name из XLSX-файла.
    Заголовки — в 1-й строке, данные — начиная со 2-й.
    Пропускает строки без ID и без Наименования проекта.
    Все заголовки нормализуются (strip), чтобы убрать trailing space
    (в исходном файле «Отделение » имеет пробел в конце).
    """
    log(f'Парсинг листа "{sheet_name}" из {xlsx_path}')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(
            f'Лист "{sheet_name}" не найден. Доступные листы: {wb.sheetnames}'
        )

    ws = wb[sheet_name]
    log(f'Размер листа: {ws.max_row} строк × {ws.max_column} колонок')

    # Читаем заголовки из 1-й строки — нормализуем (strip)
    headers = []
    for cell in ws[1]:
        val = str(cell.value).strip() if cell.value is not None else ''
        headers.append(val)
    # Убираем дубликаты пустых заголовков (только непустые)
    headers_clean = []
    for h in headers:
        if h and h not in headers_clean:
            headers_clean.append(h)
        elif not h:
            # Пропускаем пустые заголовки — но только после всех данных
            break
    log(f'Заголовки ({len(headers_clean)}): {headers_clean}')

    # Читаем данные
    projects = []
    skipped = 0
    for row_idx in range(2, ws.max_row + 1):
        # Ограничиваем чтение количеством чистых заголовков
        n_cols = len(headers_clean)
        row_values = []
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            # Обработка дат
            if isinstance(val, datetime):
                val = val.strftime('%Y-%m-%d')
            elif val is not None:
                val = str(val).strip()
                # Убираем мягкие переносы и нормализуем пробелы
                val = val.replace('\xad', '').replace('\u00a0', ' ')
                val = re.sub(r'\s+', ' ', val).strip()
            else:
                val = ''
            row_values.append(val)

        # Создаём словарь "заголовок → значение"
        record = {}
        for h, v in zip(headers_clean, row_values):
            if h:  # пропускаем пустые заголовки
                record[h] = v

        # Пропускаем строки без ID и без Наименования проекта
        id_val = record.get('ID', '').strip() if isinstance(record.get('ID'), str) else record.get('ID')
        name_val = record.get('Наименование проекта', '').strip() if isinstance(record.get('Наименование проекта'), str) else record.get('Наименование проекта')
        if (id_val == '' or id_val is None) and (name_val == '' or name_val is None):
            skipped += 1
            continue

        # Если ID — число, преобразуем
        if isinstance(id_val, str) and id_val.isdigit():
            record['ID'] = int(id_val)

        projects.append(record)

    log(f'Распарсено записей: {len(projects)}, пропущено: {skipped}')
    return projects, headers_clean


def main():
    public_key = os.environ.get('PROJECTS_PUBLIC_KEY', '').strip() or DEFAULT_PUBLIC_KEY
    sheet_name = os.environ.get('PROJECTS_SHEET_NAME', '').strip() or DEFAULT_SHEET_NAME

    try:
        # 1. Получить download_url
        download_url, filename = get_download_url(public_key)

        # 2. Скачать файл
        local_file = download_file(download_url, filename)

        # 3. Распарсить лист
        projects, headers = parse_projects(local_file, sheet_name)

        # 4. Сохранить JSON
        out = {
            'title': 'Проекты КИП пр-ва ИОС',
            'source': f'Yandex Disk (public): {public_key}',
            'sheet': sheet_name,
            'total_projects': len(projects),
            'headers': headers,
            'projects': projects,
        }
        JSON_OUT.parent.mkdir(parents=True, exist_ok=True)
        with open(JSON_OUT, 'w', encoding='utf-8') as f:
            json.dump(out, f, ensure_ascii=False, indent=2)
        log(f'JSON сохранён: {JSON_OUT}')
        log(f'Всего проектов: {len(projects)}')

        return 0

    except Exception as e:
        log(f'ОШИБКА: {e}')
        import traceback
        traceback.print_exc()
        # Если файл уже существует — не падать (используем как заглушку)
        if JSON_OUT.exists():
            log(f'Используется существующий файл: {JSON_OUT}')
            return 0
        return 1


if __name__ == '__main__':
    sys.exit(main())
